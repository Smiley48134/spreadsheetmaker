import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_sales_data(input_file, output_file):
    # Step 1: Load raw data
    if input_file.endswith('.csv'):
        df = pd.read_csv(input_file)
    elif input_file.endswith(('.xls', '.xlsx')):
        df = pd.read_excel(input_file)
    else:
        raise ValueError("Unsupported file format. Use CSV or Excel.")

    # Step 2: Drop incomplete rows
    df.dropna(subset=['Date', 'Product', 'Quantity', 'Price'], inplace=True)

    # Step 3: Convert types
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce')
    df['Price'] = pd.to_numeric(df['Price'], errors='coerce')
    df.dropna(subset=['Date', 'Quantity', 'Price'], inplace=True)

    # Step 4: Compute 'Total Sale'
    df['Total Sale'] = df['Quantity'] * df['Price']

    # Step 5: Organize columns
    df = df[['Date', 'Product', 'Quantity', 'Price', 'Total Sale']]
    df.sort_values(by='Date', inplace=True)

    # Step 6: Generate summaries
    df['Month'] = df['Date'].dt.to_period('M').astype(str)

    summary_by_product = df.groupby('Product')['Total Sale'].sum().reset_index()
    summary_by_month = df.groupby('Month')['Total Sale'].sum().reset_index()

    # Step 7: Export to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Cleaned Sales Data')
        summary_by_product.to_excel(writer, index=False, sheet_name='Sales by Product')
        summary_by_month.to_excel(writer, index=False, sheet_name='Sales by Month')

        writer.save()

    # Step 8: Apply Excel formatting
    wb = load_workbook(output_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Autofit columns (manually estimate)
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(output_file)
    print(f"Organized and summarized data saved to '{output_file}'.")

# Example usage
input_file = 'raw_sales_data.csv'   # Replace with your actual file
output_file = 'organized_sales_report.xlsx'
clean_sales_data(input_file, output_file)